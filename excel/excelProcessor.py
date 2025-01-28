import openpyxl as xl
from openpyxl.chart import Reference, BarChart


wb = xl.load_workbook("C:/Users/FAHAD/Desktop/winter'25/python/excel/products.xlsx")

sheet = wb['Sheet1']


sheet.cell(1, 4).value = "Discounted Price"


for row in range(2 , sheet.max_row + 1):
    value = sheet.cell(row , 3).value
    if(isinstance(value, str)):
        value = float(value.replace("$" , ""))

    newPrice = value * 0.9

    discountedPrices = sheet.cell(row , 4)

    discountedPrices.value = newPrice

values = Reference(sheet, min_col = 4 , min_row = 2 , max_col = 4 , max_row = sheet.max_row)

chart = BarChart()
chart.add_data(values)
sheet.add_chart(chart , "E2")

wb.save("transactionsUpdates.xlsx")



    
